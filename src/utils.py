import requests
import openpyxl
import re
import os

from src.logger import Logger
from openpyxl.styles import Font
from mimetypes import guess_extension
from datetime import datetime


class Utils:
    def download_file(url):
        try:
            response = requests.get(url)

            if not os.path.exists("output/img"):
                os.makedirs("output/img")

            if response.status_code == 200:
                filename = url.split("/")[-1]

                content_type = response.headers.get('content-type')
                extension = guess_extension(content_type.split(';')[0])

                if extension:
                    filename_with_extension = f"{filename}{extension}"
                else:
                    filename_with_extension = filename

                with open(os.path.join("output/img", filename_with_extension), "wb") as file:
                    file.write(response.content)

                return filename_with_extension
            
            Logger.warning(f"🚨 An error occurred while downloading file {url}: {response.status_code}")
            return None
        except Exception as e:
            Logger.warning(f"🚨 An error occurred while downloading file {url}: {e}")
            return None
        
    def check_date(timestamp, months):
        if months == 0:
            months = 1
            
        actual_month = datetime.now().month
        actual_year = datetime.now().year

        news_date = datetime.fromtimestamp(int(timestamp) / 1000)
        news_month = news_date.month    
        news_year = news_date.year

        month_diff = (actual_year - news_year) * 12 + (actual_month - news_month)

        if month_diff < months:
            return True
        
        return False
    
    def export_to_excel(results):
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "News Results"

            headers = ["Title", "Date", "Description", "Filename", "Count Search", "Contains Money"]

            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num, value=header)
                cell.font = Font(bold=True)

            for row_num, result in enumerate(results, 2):
                ws.cell(row=row_num, column=1, value=result["title"])
                ws.cell(row=row_num, column=2, value=result["date"])
                ws.cell(row=row_num, column=3, value=result["description"])
                ws.cell(row=row_num, column=4, value=result["filename"])
                ws.cell(row=row_num, column=5, value=result["count_search"])
                ws.cell(row=row_num, column=6, value=result["contains_money"])

            file_path = "output/news_results.xlsx"

            wb.save(file_path)
            Logger.info(f"📤✅ Results exported to {file_path}")
        except Exception as e:
            Logger.error(f"🚨 An error occurred while exporting to Excel: {e}")

    def count_search(search_text, text):
        return text.lower().count(search_text.lower())
    
    def contains_money(text):
        # Define regex pattern to match money formats
        money_pattern = r'\$[\d,]+(?:\.\d+)?(?:\s*(?:USD|dollars?))?'

        # Search for money pattern in the text
        if re.search(money_pattern, text):
            return True
        else:
            return False
        